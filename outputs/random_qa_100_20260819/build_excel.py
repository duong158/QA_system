from __future__ import annotations

import json
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
SOURCE = Path(os.getenv("QA_RESULTS_JSON", str(ROOT / "combined_results.json"))).resolve()
OUTPUT = Path(
    os.getenv("QA_EXCEL_OUTPUT", str(ROOT / "ket_qua_100_cau_hoi_ngau_nhien.xlsx"))
).resolve()

NAVY = "17365D"
BLUE = "1F4E78"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF3F8"
GREEN = "E2F0D9"
GREEN_TEXT = "217346"
RED = "FCE4D6"
RED_TEXT = "C00000"
GOLD = "FFF2CC"
WHITE = "FFFFFF"
TEXT = "243447"
MUTED = "667085"
GRID = "D8E2EA"


def result_note(row: dict) -> str:
    gold = str(row.get("gold_answer") or "").strip()
    prediction = str(row.get("predicted_answer") or "").strip()
    if int(row.get("exact_match", 0)) == 1:
        return "Khớp chính xác" if gold else "Đúng khi không trả lời"
    if gold and not prediction:
        return "Không trả lời"
    if not gold and prediction:
        return "Trả lời khi câu không có đáp án"
    return "Không khớp đáp án chuẩn"


payload = json.loads(SOURCE.read_text(encoding="utf-8"))
rows = payload["predictions"]
if len(rows) != 100:
    raise ValueError(f"Expected 100 predictions, found {len(rows)}")
reader_name = str(payload.get("reader") or "unknown")
retriever_name = str(payload.get("retriever") or "unknown")
checkpoint_name = Path(str(payload.get("checkpoint") or "unknown")).name
dense_model = str(payload.get("dense_model") or "n/a")
top_k = int(payload.get("top_k") or 10)
expected_correct = sum(int(item.get("exact_match", 0)) for item in rows)
expected_incorrect = len(rows) - expected_correct

wb = Workbook()
ws = wb.active
ws.title = "Kết quả QA"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 85
ws.freeze_panes = "A9"

wb.creator = "Codex"
wb.title = f"Kết quả 100 câu - {reader_name} + {retriever_name}"
wb.subject = "Đánh giá hệ thống QA bằng Exact Match"
wb.description = (
    f"100 câu ViQuAD validation, seed 20260819; reader={reader_name}, "
    f"checkpoint={checkpoint_name}, retriever={retriever_name}."
)
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True

ws.merge_cells("A1:H1")
ws["A1"] = "KẾT QUẢ CHẠY THỬ 100 CÂU HỎI NGẪU NHIÊN"
ws["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:H2")
ws["A2"] = (
    f"Reader: {reader_name} ({checkpoint_name})  |  Retriever: {retriever_name}  |  "
    f"Dense: {dense_model}  |  top-k {top_k}  |  Seed: 20260819"
)
ws["A2"].font = Font(name="Aptos", size=10, italic=True, color=MUTED)
ws["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 34

cards = [
    ("A4:B4", "A5:B5", "TỔNG SỐ CÂU", "=COUNTA(A9:A108)", LIGHT_BLUE),
    ("C4:D4", "C5:D5", "ĐÚNG (EXACT MATCH)", '=COUNTIF(F9:F108,"✓")', GREEN),
    ("E4:F4", "E5:F5", "SAI", '=COUNTIF(F9:F108,"✗")', RED),
    ("G4:H4", "G5:H5", "TỶ LỆ ĐÚNG", "=C5/A5", GOLD),
]
for label_range, value_range, label, formula, fill_color in cards:
    ws.merge_cells(label_range)
    ws.merge_cells(value_range)
    label_cell = ws[label_range.split(":")[0]]
    value_cell = ws[value_range.split(":")[0]]
    label_cell.value = label
    value_cell.value = formula
    for rng in (label_range, value_range):
        for row_cells in ws[rng]:
            for cell in row_cells:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.font = Font(name="Aptos", size=9, bold=True, color=TEXT)
    value_cell.font = Font(name="Aptos Display", size=16, bold=True, color=TEXT)

ws["G5"].number_format = "0.0%"
ws.row_dimensions[4].height = 22
ws.row_dimensions[5].height = 31

ws.merge_cells("A7:H7")
ws["A7"] = (
    "Quy tắc chấm: ĐÚNG khi đáp án hệ thống khớp Exact Match sau chuẩn hóa "
    "(không phân biệt hoa/thường, dấu câu và dấu gạch dưới); ngược lại là SAI."
)
ws["A7"].font = Font(name="Aptos", size=9, italic=True, color=MUTED)
ws["A7"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[7].height = 30

headers = [
    "STT",
    "Mã câu",
    "Câu hỏi",
    "Đáp án hệ thống",
    "Đáp án chuẩn",
    "Đúng?",
    "F1",
    "Ghi chú",
]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=8, column=col, value=header)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[8].height = 28

thin_bottom = Border(bottom=Side(style="thin", color=GRID))
for index, item in enumerate(rows, start=1):
    excel_row = index + 8
    is_correct = int(item.get("exact_match", 0)) == 1
    values = [
        index,
        str(item.get("id") or ""),
        str(item.get("question") or ""),
        str(item.get("predicted_answer") or ""),
        str(item.get("gold_answer") or ""),
        "✓" if is_correct else "✗",
        float(item.get("f1", 0.0)),
        result_note(item),
    ]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=excel_row, column=col, value=value)
        cell.font = Font(name="Aptos", size=10, color=TEXT)
        cell.border = thin_bottom
        cell.alignment = Alignment(
            horizontal="center" if col in (1, 2, 6, 7) else "left",
            vertical="top",
            wrap_text=col in (3, 4, 5, 8),
        )
    ws.cell(excel_row, 6).font = Font(
        name="Segoe UI Symbol", size=14, bold=True, color=GREEN_TEXT if is_correct else RED_TEXT
    )
    ws.cell(excel_row, 7).number_format = "0.0%"
    wrapped_lines = max(
        1,
        (len(str(values[2])) + 64) // 65,
        (len(str(values[3])) + 43) // 44,
        (len(str(values[4])) + 43) // 44,
        (len(str(values[7])) + 34) // 35,
    )
    ws.row_dimensions[excel_row].height = min(90, max(30, 10 + wrapped_lines * 14))

table = Table(displayName="KetQuaQA100", ref="A8:H108")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

green_fill = PatternFill("solid", fgColor=GREEN)
red_fill = PatternFill("solid", fgColor=RED)
ws.conditional_formatting.add(
    "F9:F108",
    CellIsRule(operator="equal", formula=['"✓"'], fill=green_fill),
)
ws.conditional_formatting.add(
    "F9:F108",
    CellIsRule(operator="equal", formula=['"✗"'], fill=red_fill),
)

widths = {
    "A": 7,
    "B": 18,
    "C": 53,
    "D": 38,
    "E": 38,
    "F": 10,
    "G": 10,
    "H": 31,
}
for column, width in widths.items():
    ws.column_dimensions[column].width = width

ws.auto_filter.ref = "A8:H108"
ws.print_title_rows = "8:8"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.sheet_properties.outlinePr.summaryBelow = True
ws.page_margins.left = 0.25
ws.page_margins.right = 0.25
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5

# Technical audit sheet retains raw evaluator fields without cluttering the main table.
raw = wb.create_sheet("Chi tiết kỹ thuật")
raw.sheet_view.showGridLines = False
raw.freeze_panes = "A2"
raw.sheet_view.zoomScale = 90
raw_headers = [
    "STT", "ID", "Câu hỏi", "Đáp án chuẩn", "Đáp án hệ thống", "Exact Match",
    "F1", "Có đáp án", "Loại câu hỏi", "Phương pháp", "Retrieval hit",
    "Reader score", "Ranking score", "Answer-type score", "Độ trễ (ms)",
    "Retriever", "Reader", "Checkpoint",
]
raw.append(raw_headers)
for index, item in enumerate(rows, start=1):
    raw.append([
        index,
        str(item.get("id") or ""),
        str(item.get("question") or ""),
        str(item.get("gold_answer") or ""),
        str(item.get("predicted_answer") or ""),
        int(item.get("exact_match", 0)),
        float(item.get("f1", 0.0)),
        bool(item.get("is_answerable", False)),
        str(item.get("question_type") or ""),
        str(item.get("reader_method") or ""),
        item.get("retrieval_hit"),
        item.get("reader_score"),
        item.get("ranking_score"),
        item.get("answer_type_score"),
        float(item.get("latency_ms", 0.0)),
        str(item.get("retriever") or retriever_name),
        str(item.get("reader") or reader_name),
        str(item.get("checkpoint") or payload.get("checkpoint") or ""),
    ])

for cell in raw[1]:
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
raw.row_dimensions[1].height = 42
for row_index, row_cells in enumerate(raw.iter_rows(min_row=2, max_row=101), start=2):
    for cell in row_cells:
        cell.font = Font(name="Aptos", size=9, color=TEXT)
        cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (3, 4, 5))
        cell.border = thin_bottom
    wrapped_lines = max(
        1,
        (len(str(raw.cell(row_index, 3).value or "")) + 63) // 64,
        (len(str(raw.cell(row_index, 4).value or "")) + 41) // 42,
        (len(str(raw.cell(row_index, 5).value or "")) + 41) // 42,
    )
    raw.row_dimensions[row_index].height = min(72, max(24, 8 + wrapped_lines * 13))
raw.column_dimensions["A"].width = 7
raw.column_dimensions["B"].width = 18
raw.column_dimensions["C"].width = 52
raw.column_dimensions["D"].width = 34
raw.column_dimensions["E"].width = 34
for column in "FGHIJKLMNO":
    raw.column_dimensions[column].width = 16
raw.column_dimensions["P"].width = 14
raw.column_dimensions["Q"].width = 12
raw.column_dimensions["R"].width = 48
for row_index in range(2, 102):
    raw.cell(row_index, 7).number_format = "0.0%"
    for col_index in (12, 13, 14):
        raw.cell(row_index, col_index).number_format = "0.000"
    raw.cell(row_index, 15).number_format = "#,##0"
raw_table = Table(displayName="ChiTietQA100", ref="A1:R101")
raw_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
raw.add_table(raw_table)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)

# Compact structural verification after re-opening the produced workbook.
check = load_workbook(OUTPUT, data_only=False)
main = check["Kết quả QA"]
technical = check["Chi tiết kỹ thuật"]
assert main.max_row == 108 and main.max_column == 8
assert technical.max_row == 101 and technical.max_column == 18
assert len(main.tables) == 1 and len(technical.tables) == 1
assert main["A5"].value == "=COUNTA(A9:A108)"
assert main["C5"].value == '=COUNTIF(F9:F108,"✓")'
assert sum(main.cell(row, 6).value == "✓" for row in range(9, 109)) == expected_correct
assert sum(main.cell(row, 6).value == "✗" for row in range(9, 109)) == expected_incorrect
assert all(technical.cell(row, 16).value == retriever_name for row in range(2, 102))
assert all(technical.cell(row, 17).value == reader_name for row in range(2, 102))
for sheet in check.worksheets:
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            if isinstance(cell.value, str) and any(
                error in cell.value
                for error in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
            ):
                raise ValueError(f"Formula error literal at {sheet.title}!{cell.coordinate}")

print(
    json.dumps(
        {
            "output": str(OUTPUT),
            "rows": 100,
            "correct": expected_correct,
            "incorrect": expected_incorrect,
            "reader": reader_name,
            "retriever": retriever_name,
            "checkpoint": checkpoint_name,
            "sheets": check.sheetnames,
            "formula_cells": {
                "total": main["A5"].value,
                "correct": main["C5"].value,
                "incorrect": main["E5"].value,
                "accuracy": main["G5"].value,
            },
        },
        ensure_ascii=True,
        indent=2,
    )
)
