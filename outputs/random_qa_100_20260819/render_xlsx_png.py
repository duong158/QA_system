from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


SCRIPT_ROOT = Path(__file__).resolve().parent
SOURCE = Path(
    os.getenv(
        "QA_EXCEL_SOURCE",
        str(SCRIPT_ROOT / "ket_qua_100_cau_hoi_ngau_nhien.xlsx"),
    )
).resolve()
ROOT = SOURCE.parent
FONT_REGULAR = Path(r"C:\Windows\Fonts\arial.ttf")
FONT_BOLD = Path(r"C:\Windows\Fonts\arialbd.ttf")
FONT_SYMBOL = Path(r"C:\Windows\Fonts\seguisym.ttf")


def to_hex(color, default: str) -> str:
    if color is None:
        return default
    if color.type == "rgb" and color.rgb:
        return f"#{color.rgb[-6:]}"
    if color.type == "indexed" and color.indexed is not None:
        return f"#{COLOR_INDEX[color.indexed][-6:]}"
    return default


def font_for(cell):
    size = max(10, int(round(float(cell.font.sz or 10) * 1.25)))
    if cell.font.name == "Segoe UI Symbol":
        path = FONT_SYMBOL
    else:
        path = FONT_BOLD if cell.font.bold else FONT_REGULAR
    return ImageFont.truetype(str(path), size=size)


def fit_lines(draw, text: str, font, max_width: int, max_height: int):
    words = str(text or "").split()
    if not words:
        return []
    lines = []
    line = words[0]
    for word in words[1:]:
        candidate = f"{line} {word}"
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            line = candidate
        else:
            lines.append(line)
            line = word
    lines.append(line)
    line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + 3
    max_lines = max(1, max_height // line_height)
    clipped = len(lines) > max_lines
    lines = lines[:max_lines]
    if clipped:
        last = lines[-1]
        while last and draw.textbbox((0, 0), last + "…", font=font)[2] > max_width:
            last = last[:-1]
        lines[-1] = last.rstrip() + "…"
    return lines


def render_sheet(ws, out_path: Path, min_row: int, max_row: int, min_col: int, max_col: int, replacements=None):
    replacements = replacements or {}
    col_widths = []
    for col in range(min_col, max_col + 1):
        width = ws.column_dimensions[get_column_letter(col)].width or 12
        col_widths.append(max(55, int(width * 7.2)))
    row_heights = []
    for row in range(min_row, max_row + 1):
        height = ws.row_dimensions[row].height or 22
        row_heights.append(max(24, int(height * 1.33)))

    margin_x, margin_y = 18, 48
    image = Image.new("RGB", (sum(col_widths) + margin_x * 2, sum(row_heights) + margin_y + 18), "#EEF3F7")
    draw = ImageDraw.Draw(image)
    heading_font = ImageFont.truetype(str(FONT_BOLD), 20)
    draw.text((margin_x, 14), f"Sheet: {ws.title}", fill="#17365D", font=heading_font)

    x_positions = [margin_x]
    for width in col_widths:
        x_positions.append(x_positions[-1] + width)
    y_positions = [margin_y]
    for height in row_heights:
        y_positions.append(y_positions[-1] + height)

    merged = {}
    covered = set()
    for area in ws.merged_cells.ranges:
        if area.max_row < min_row or area.min_row > max_row or area.max_col < min_col or area.min_col > max_col:
            continue
        anchor = (area.min_row, area.min_col)
        merged[anchor] = (area.max_row, area.max_col)
        for row in range(area.min_row, area.max_row + 1):
            for col in range(area.min_col, area.max_col + 1):
                if (row, col) != anchor:
                    covered.add((row, col))

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in covered:
                continue
            cell = ws.cell(row, col)
            end_row, end_col = merged.get((row, col), (row, col))
            x0 = x_positions[col - min_col]
            x1 = x_positions[min(end_col, max_col) - min_col + 1]
            y0 = y_positions[row - min_row]
            y1 = y_positions[min(end_row, max_row) - min_row + 1]
            fill = to_hex(cell.fill.fgColor, "#FFFFFF") if cell.fill.fill_type else "#FFFFFF"
            if col == 6 and row >= 9 and cell.value == "✓":
                fill = "#E2F0D9"
            elif col == 6 and row >= 9 and cell.value == "✗":
                fill = "#FCE4D6"
            draw.rectangle((x0, y0, x1, y1), fill=fill, outline="#D8E2EA", width=1)

            value = replacements.get(cell.coordinate, cell.value)
            if value is None:
                continue
            if isinstance(value, float) and cell.number_format and "%" in cell.number_format:
                text = f"{value:.1%}"
            else:
                text = str(value)
            font = font_for(cell)
            text_color = to_hex(cell.font.color, "#243447")
            lines = fit_lines(draw, text, font, max(8, x1 - x0 - 14), max(8, y1 - y0 - 10))
            if not lines:
                continue
            line_height = draw.textbbox((0, 0), "Ag", font=font)[3] + 3
            block_height = len(lines) * line_height
            vertical = cell.alignment.vertical or "top"
            if vertical == "center":
                text_y = y0 + max(5, (y1 - y0 - block_height) // 2)
            elif vertical == "bottom":
                text_y = y1 - block_height - 5
            else:
                text_y = y0 + 5
            for line in lines:
                line_width = draw.textbbox((0, 0), line, font=font)[2]
                horizontal = cell.alignment.horizontal or "left"
                if horizontal == "center":
                    text_x = x0 + max(6, (x1 - x0 - line_width) // 2)
                elif horizontal == "right":
                    text_x = x1 - line_width - 7
                else:
                    text_x = x0 + 7
                draw.text((text_x, text_y), line, fill=text_color, font=font)
                text_y += line_height

    image.save(out_path)
    return image.size


wb = load_workbook(SOURCE, data_only=False)
main_sheet = wb["Kết quả QA"]
technical_sheet = wb["Chi tiết kỹ thuật"]
correct_count = sum(main_sheet.cell(row, 6).value == "✓" for row in range(9, 109))
incorrect_count = 100 - correct_count
preview_suffix = SOURCE.stem
main_size = render_sheet(
    main_sheet,
    ROOT / f"preview_ket_qua_{preview_suffix}.png",
    1,
    18,
    1,
    8,
    replacements={
        "A5": "100",
        "C5": str(correct_count),
        "E5": str(incorrect_count),
        "G5": f"{correct_count:.1f}%",
    },
)
detail_size = render_sheet(
    technical_sheet,
    ROOT / f"preview_chi_tiet_{preview_suffix}.png",
    1,
    12,
    1,
    technical_sheet.max_column,
)
first_correct_row = next(
    row for row in range(9, 109) if main_sheet.cell(row, 6).value == "✓"
)
checkmark_size = render_sheet(
    main_sheet,
    ROOT / f"preview_checkmark_{preview_suffix}.png",
    max(9, first_correct_row - 2),
    min(108, first_correct_row + 2),
    1,
    8,
)
print({"main": main_size, "detail": detail_size, "checkmark": checkmark_size})
