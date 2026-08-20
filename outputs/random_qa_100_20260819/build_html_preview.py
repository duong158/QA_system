from __future__ import annotations

import html
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "ket_qua_100_cau_hoi_ngau_nhien.xlsx"
OUTPUT = ROOT / "workbook_preview.html"


def color_value(color, default: str) -> str:
    if color is None:
        return default
    if color.type == "rgb" and color.rgb:
        return f"#{color.rgb[-6:]}"
    if color.type == "indexed" and color.indexed is not None:
        value = COLOR_INDEX[color.indexed]
        return f"#{value[-6:]}"
    return default


def display_value(cell, replacements: dict[str, str]) -> str:
    value = replacements.get(cell.coordinate, cell.value)
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if cell.number_format and "%" in cell.number_format:
            return f"{value:.1%}"
        return f"{value:.3f}".rstrip("0").rstrip(".")
    return str(value)


def render_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, replacements=None):
    replacements = replacements or {}
    merged_anchors = {}
    covered = set()
    for merged in ws.merged_cells.ranges:
        if merged.max_row < min_row or merged.min_row > max_row:
            continue
        if merged.max_col < min_col or merged.min_col > max_col:
            continue
        anchor = (merged.min_row, merged.min_col)
        merged_anchors[anchor] = (merged.max_col - merged.min_col + 1, merged.max_row - merged.min_row + 1)
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if (row, col) != anchor:
                    covered.add((row, col))

    colgroup = []
    for col in range(min_col, max_col + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width or 12
        colgroup.append(f'<col style="width:{max(55, width * 7.2):.0f}px">')

    parts = ["<table><colgroup>", *colgroup, "</colgroup>"]
    for row in range(min_row, max_row + 1):
        height = ws.row_dimensions[row].height or 22
        parts.append(f'<tr style="height:{height * 1.33:.0f}px">')
        for col in range(min_col, max_col + 1):
            if (row, col) in covered:
                continue
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            colspan, rowspan = merged_anchors.get((row, col), (1, 1))
            fill = color_value(cell.fill.fgColor, "#FFFFFF") if cell.fill.fill_type else "#FFFFFF"
            font_color = color_value(cell.font.color, "#243447")
            horizontal = cell.alignment.horizontal or "left"
            vertical = cell.alignment.vertical or "top"
            weight = "700" if cell.font.bold else "400"
            style = (
                f"background:{fill};color:{font_color};font-weight:{weight};"
                f"font-size:{cell.font.sz or 10}pt;text-align:{horizontal};vertical-align:{vertical};"
            )
            attrs = f' colspan="{colspan}" rowspan="{rowspan}"' if (colspan, rowspan) != (1, 1) else ""
            value = html.escape(display_value(cell, replacements))
            parts.append(f'<td{attrs} style="{style}">{value}</td>')
        parts.append("</tr>")
    parts.append("</table>")
    return "".join(parts)


wb = load_workbook(SOURCE, data_only=False)
main = wb["Kết quả QA"]
detail = wb["Chi tiết kỹ thuật"]

main_html = render_range(
    main,
    1,
    18,
    1,
    8,
    replacements={"A5": "100", "C5": "8", "E5": "92", "G5": "8.0%"},
)
detail_html = render_range(detail, 1, 12, 1, 15)

document = f"""<!doctype html>
<html lang="vi">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Xem trước workbook QA</title>
<style>
  body {{ margin:0; padding:24px; background:#eef3f7; color:#243447; font-family:Aptos, Arial, sans-serif; }}
  h2 {{ margin:0 0 12px; font-size:18px; color:#17365D; }}
  .sheet {{ background:white; border-radius:10px; box-shadow:0 4px 18px rgba(23,54,93,.12); padding:18px; margin-bottom:26px; overflow:auto; }}
  table {{ border-collapse:collapse; table-layout:fixed; min-width:max-content; }}
  td {{ border-bottom:1px solid #D8E2EA; padding:7px 9px; white-space:normal; overflow-wrap:anywhere; line-height:1.25; }}
</style>
</head>
<body>
  <section class="sheet"><h2>Sheet: Kết quả QA</h2>{main_html}</section>
  <section class="sheet"><h2>Sheet: Chi tiết kỹ thuật</h2>{detail_html}</section>
</body>
</html>"""

OUTPUT.write_text(document, encoding="utf-8")
print(str(OUTPUT))
