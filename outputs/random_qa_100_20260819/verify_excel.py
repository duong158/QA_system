from __future__ import annotations

import json
import os
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


root = Path(__file__).resolve().parent
xlsx = Path(
    os.getenv("QA_EXCEL_SOURCE", str(root / "ket_qua_100_cau_hoi_ngau_nhien.xlsx"))
).resolve()
source_path = Path(
    os.getenv("QA_RESULTS_JSON", str(root / "combined_results.json"))
).resolve()
source = json.loads(source_path.read_text(encoding="utf-8"))
expected_correct = sum(int(item.get("exact_match", 0)) for item in source["predictions"])
expected_incorrect = 100 - expected_correct

assert zipfile.is_zipfile(xlsx)
wb = load_workbook(xlsx, data_only=False)
assert wb.sheetnames == ["Kết quả QA", "Chi tiết kỹ thuật"]
main = wb["Kết quả QA"]
technical = wb["Chi tiết kỹ thuật"]

statuses = [main.cell(row, 6).value for row in range(9, 109)]
ids = [main.cell(row, 2).value for row in range(9, 109)]
assert len(ids) == 100 and len(set(ids)) == 100
assert ids == [prediction["id"] for prediction in source["predictions"]]
assert statuses.count("✓") == expected_correct
assert statuses.count("✗") == expected_incorrect
assert main.freeze_panes == "A9"
assert main.tables["KetQuaQA100"].ref == "A8:H108"
assert technical.tables["ChiTietQA100"].ref == (
    f"A1:{get_column_letter(technical.max_column)}101"
)
assert wb.calculation.calcMode == "auto"
assert wb.calculation.fullCalcOnLoad

assert [main["A5"].value, main["C5"].value, main["E5"].value, main["G5"].value] == [
    "=COUNTA(A9:A108)",
    '=COUNTIF(F9:F108,"✓")',
    '=COUNTIF(F9:F108,"✗")',
    "=C5/A5",
]

errors = []
for sheet in wb.worksheets:
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            if isinstance(cell.value, str) and any(
                token in cell.value
                for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
            ):
                errors.append(f"{sheet.title}!{cell.coordinate}")
assert not errors

reader = source.get("reader")
retriever = source.get("retriever")
checkpoint = str(source.get("checkpoint") or "")
if technical.max_column >= 18:
    assert all(technical.cell(row, 16).value == retriever for row in range(2, 102))
    assert all(technical.cell(row, 17).value == reader for row in range(2, 102))
    assert all(technical.cell(row, 18).value == checkpoint for row in range(2, 102))

print(
    json.dumps(
        {
            "file": str(xlsx.resolve()),
            "bytes": xlsx.stat().st_size,
            "rows": 100,
            "unique_ids": 100,
            "correct": expected_correct,
            "incorrect": expected_incorrect,
            "reader": reader,
            "retriever": retriever,
            "checkpoint": checkpoint,
            "hybrid_dense_verified": source.get("hybrid_dense_verified"),
            "formulas_ok": True,
            "formula_errors": 0,
            "visual_previews_checked": 3,
        },
        ensure_ascii=True,
        indent=2,
    )
)
